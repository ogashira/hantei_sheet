import platform
import os
import openpyxl
import pandas as pd
import subprocess
from typing import List
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.worksheet.worksheet import Worksheet
from kensa_hinban import KensaHinban


class Excel:
    '''
    1ページのレイアウト
    先頭２行：タイトルと生産予定日  ->  ROW_TITLE
    １ブロック: １０行              ->  ROW_IN_BLOCK
    1ページに４ブロック表示         ->  BLOCKS_IN_PAGE
    ブロックとブロックの間は１行
    １ページの行数は 2 + 10 * 4 + (4-1) = 45行
    '''

    def __init__(self, kensaHinbans: List[KensaHinban], sz_yt_date: str)->None:
        self.__pf:str = platform.system()
        path = './'
        if self.__pf == 'Windows':
            path = r'//192.168.1.247/共有/技術課ﾌｫﾙﾀﾞ/200. effit_data/ﾏｽﾀ/hantei_format.xlsx'
        if self.__pf == 'Linux':
            path = r'/mnt/public/技術課ﾌｫﾙﾀﾞ/200. effit_data/ﾏｽﾀ/hantei_format.xlsx'
        if self.__pf == 'Darwin':
            path = r'/Volumes/共有/技術課ﾌｫﾙﾀﾞ/200. effit_data/ﾏｽﾀ/hantei_format.xlsx'

        self.__kensaHinbans: List[KensaHinban] = kensaHinbans
        self.__sz_yt_date: str = sz_yt_date

        self.wb:openpyxl.Workbook = openpyxl.load_workbook(filename= path,
                                                                 data_only=True)
        self.ws:Worksheet = self.wb['format']


        self.ROW_TITLE: int = 2       # ページの先頭タイトルの行数
        self.ROW_IN_BLOCK: int = 10   # 1つの品番のデータの行数
        self.BLOCKS_IN_PAGE: int = 4  # 1ページに表示する品番数

        self.__pages:int = self.calc_pages() 
        self.__lastRow = self.calc_lastRow()

        self.delete_rows()
        self.filling_date()
        self.filling_yotei()

        self.print_setting()

        self.save_file()

        self.open_file()


    def calc_pages(self)-> int:
        count_data: int = len(self.__kensaHinbans)
        syou: int = count_data // self.BLOCKS_IN_PAGE
        amari: int = count_data % self.BLOCKS_IN_PAGE
        pages: int = syou
        if amari > 0:
            pages = syou + 1

        return pages


    def calc_lastRow(self)-> int:
        '''
        blocks_count == self.__kensaHinbansのデータ数 から最終行を求める
        '''
        blocks_count: int = len(self.__kensaHinbans)
        blocks_count_lastPage: int = blocks_count % self.BLOCKS_IN_PAGE  # 余り
        if blocks_count_lastPage == 0:
            blocks_count_lastPage = 4
        beforeLastPage_lastRow: int = ((self.ROW_TITLE 
                                    + self.BLOCKS_IN_PAGE * self.ROW_IN_BLOCK 
                                    + self.BLOCKS_IN_PAGE -1) 
                                    * (self.__pages-1)) 
        lastPage_lastRow: int =  (self.ROW_TITLE 
                                  + blocks_count_lastPage * self.ROW_IN_BLOCK 
                                  + blocks_count_lastPage -1)
        if blocks_count_lastPage == 0:
            lastPage_lastRow = 0
        lastRow: int = beforeLastPage_lastRow + lastPage_lastRow

        return lastRow


    def delete_rows(self)-> None:
        self.ws.delete_rows(self.__lastRow + 1, 450)


    def filling_date(self) -> None:

        row_of_one_page = (self.ROW_TITLE 
                           + self.ROW_IN_BLOCK * self.BLOCKS_IN_PAGE 
                           + self.BLOCKS_IN_PAGE - 1 )
        for i in range(self.__pages):
            put_row = row_of_one_page * i + 1
            self.ws.cell(row=put_row,     column=5).value = \
                                       self.__sz_yt_date # type: ignore
            # mergeセルにデータを入れようとしてない？って
            # type: ignoreを入れないとPyrightが警告出す。


    def filling_yotei(self)-> None:
        '''
        KensaHinbanのインスタンスメソッドを呼び出して、excelに予定データ
        を入力する。
        何行目から入力するのかは。calc_firstRow_for_blockにkensaHinbans
        のindexを渡して計算する
        '''

        def calc_firstRow_for_block(blocks_count: int)-> int:
            '''
            self.__kensaHinbansの順番から何行目からデータを入れるのかを求める
            '''
            pages: int = blocks_count // self.BLOCKS_IN_PAGE          # 商 1ページの場合は0
            blocks_count_lastPage: int = blocks_count % self.BLOCKS_IN_PAGE  # 余り
            row_of_one_page = (self.ROW_TITLE 
                               + self.BLOCKS_IN_PAGE * self.ROW_IN_BLOCK 
                               + self.BLOCKS_IN_PAGE - 1)
            firstRow_for_block = (row_of_one_page * pages 
                                  + blocks_count_lastPage * self.ROW_IN_BLOCK
                                  + blocks_count_lastPage
                                  + self.ROW_TITLE
                                  + 1    # ～の次
                                  )
            return firstRow_for_block
        
        for index, kensaHinban in enumerate(self.__kensaHinbans):

            # データ入力の基準(先頭)の行を取得する
            standerd_row: int = calc_firstRow_for_block(index)

            #kensaHinbanインスタンスのfilling_dataメソッド呼ぶ
            kensaHinban.filling_data(self.ws, standerd_row)


    def print_setting(self)-> None:
          
        # print設定
        CM:float = 1 / 2.54 # センチメートル 使っていない
        max_row:int = self.__lastRow
        print_area:str = 'A1:H' + str(max_row)
        self.ws.print_area = print_area 


    def save_file(self)->None:
        try:
            self.wb.save(filename= f'./record_sheet.xlsx')
        except Exception:
            print('ファイルを保存できません。record_sheet.xlsxを閉じてください')

    def open_file(self)-> None:
        
        openFile: str = r'record_sheet.xlsx'

        if platform.system() == 'Windows':
            try:
                os.startfile(openFile)
            except OSError as e:
                print(f'./record_sheet.xlsxを開けませんでした{e}')
        if platform.system() == 'Darwin':
            try:
                subprocess.run(['open', openFile], check=True)
            except (OSError, subprocess.CalledProcessError):
                print('./record_sheet.xlsxを開けませんでした')
        if platform.system() == 'Linux':
            print('手動で./record_sheet.xlsxを開いてください')






